from django.shortcuts import render
from django.http import HttpResponse
import csv

from djangoldp.models import Model
from djangoldp.views import NoCSRFAuthentication
from rest_framework.response import Response 
from rest_framework.views import APIView
from django.http import HttpResponseNotFound

from openpyxl import load_workbook
import validators
import io
import os.path

PACKAGE_PATH = os.path.realpath(os.path.dirname(__file__))

class XLSXExportView(APIView):
    # authentication_classes = (NoCSRFAuthentication,)

    def dispatch(self, request, *args, **kwargs):
        '''overriden dispatch method to append some custom headers'''
        response = super(XLSXExportView, self).dispatch(request, *args, **kwargs)
        response["Access-Control-Allow-Origin"] = request.META.get('HTTP_ORIGIN')
        response["Access-Control-Allow-Methods"] = "POST"
        response["Access-Control-Allow-Headers"] = "authorization, Content-Type, if-match, accept, cache-control, pragma, user-agent"
        response["Access-Control-Expose-Headers"] = "Location, User"
        response["Access-Control-Allow-Credentials"] = 'true'
        response["Accept-Post"] = "*/*"
        response["Accept"] = "*/*"
        return response

    def post(self, request):
        # Check that we get an array
        if (request.method == 'POST' and request.data and isinstance(request.data, list)):

            # Open the template file located in templates/export/
            templateFile = load_workbook(os.path.join(PACKAGE_PATH, 'templates/export/template.xlsx'))
            destFile = 'export_societes.xlsx'
            activeWorksheet = templateFile.active

            # Just make sure you can start writing at line 8 column 2
            rowNumber = 8
            columnNumber = 2

            for urlid in request.data:
              # Check that the array entries are URLs
              if validators.url(urlid):
                # Check that the corresponding Actors exists
                model, instance = Model.resolve(urlid)
                if instance and instance.activity:
                  # Generate XLSX file using the lib, see the lib readme
                  if instance.siren:
                    siren = instance.siren
                  else:
                    siren = "N/A"
                  
                  if instance.website:
                    website = instance.website
                  else:
                    website = "N/A"

                  if instance.mail:
                    mail = instance.mail
                  else:
                    mail = "N/A"

                  if instance.phone:
                    phone = instance.phone
                  else:
                    phone = "N/A"

                  if instance.logo:
                    logo = instance.logo
                  else:
                    logo = "N/A"

                  activeWorksheet.cell(rowNumber, columnNumber, value=instance.name)
                  activeWorksheet.cell(rowNumber, columnNumber + 1, value=instance.activity.name)
                  activeWorksheet.cell(rowNumber, columnNumber + 2, value=siren)
                  activeWorksheet.cell(rowNumber, columnNumber + 3, value=instance.address)
                  activeWorksheet.cell(rowNumber, columnNumber + 4, value=instance.postcode)
                  activeWorksheet.cell(rowNumber, columnNumber + 5, value=instance.city)
                  activeWorksheet.cell(rowNumber, columnNumber + 6, value=website)
                  activeWorksheet.cell(rowNumber, columnNumber + 7, value=mail)
                  activeWorksheet.cell(rowNumber, columnNumber + 8, value=phone)
                  activeWorksheet.cell(rowNumber, columnNumber + 9, value=logo)
                  rowNumber += 1
            
            templateFile.save(destFile)
            try:
                from django.http import FileResponse
                response = FileResponse(io.open(destFile, mode="rb"))
                response['Content-Disposition'] = 'attachment; filename=' + destFile
                return response
            except IOError:
                # handle the case where file does not exist
                response = HttpResponseNotFound('<h1>File does not exist</h1>')

        return Response(status=204)

