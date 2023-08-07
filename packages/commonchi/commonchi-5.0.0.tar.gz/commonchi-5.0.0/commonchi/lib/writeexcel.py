#!/usr/bin/env python
# _*_ coding:utf-8 _*_
from commonchi.lib.utility import report_path

__author__ = 'jack'

import os, sys
from time import sleep

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
# from cases.ussm2 import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, DARKYELLOW
# import configparser as cparser
#
# # --------- 读取config.ini配置文件 ---------------
# cf = cparser.ConfigParser()
# cf.read(setting.TEST_CONFIG, encoding='UTF-8')
# name = cf.get("tester", "name")


class WriteExcel():
    """文件写入数据"""

    def __init__(self, fileName,source):
        self.filename = fileName
        self.source = source

        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            # shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)
            shutil.copyfile(self.source,self.filename)
            sleep(1)

        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active


    def write_data(self, row_n, value, name):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :param value: 运行人员
        :return: 无
        """
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
        self.ws.cell(row_n, 13, name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)


    #cxx2022.1.22
    def write_data_1(self, row_n, value, name,time_1,actual_results):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :param value: 运行人员
        :return: 无
        """

        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        K_n = "AF" + str(row_n)
        L_n = "AG" + str(row_n)
        M_n = "AH" + str(row_n)
        N_n = "AI" + str(row_n)
        #实际结果
        self.ws.cell(row_n, 32, actual_results)
        self.ws[N_n].font = font1
        self.ws[N_n].alignment = align
        #PASS or FAIL
        if value == "PASS":
            self.ws.cell(row_n, 33, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 33, value)
            self.ws[L_n].font = font_RED
        self.ws.cell(row_n, 34, name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        #time
        self.ws.cell(row_n, 35, time_1)
        self.ws[N_n].font = font1
        self.ws[N_n].alignment = align
        self.wb.save(self.filename)
