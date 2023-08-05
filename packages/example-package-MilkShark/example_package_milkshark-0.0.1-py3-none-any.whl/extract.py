import pyautogui as gui
import time
import os
import sys
import xlrd  #读取xls
import openpyxl #excel 不能读取xls
from openpyxl import Workbook
from pynput.keyboard import Controller
import win32gui


#去除掉不要打卡的名字
def remove_dup(contents):
    print(len(contents))
    tmp_dic = {content[0]:content for content in contents}
    new_contents = tmp_dic.values()
    print(len(new_contents))

    return new_contents
def write_excel(contents,save_path):
    wb = Workbook()
    excel_head = ['学号','姓名','电话','学历']
    sheet = wb.create_sheet("sheet",0)
    for i,content in enumerate(contents):
        for j,item in enumerate(content):
            print(item)
            sheet.cell(row = i+1,column=j+1,value=item)
    wb.save(save_path)


def get_excel_path(base_path):
    dirs = os.listdir(base_path)
    full_paths = [os.path.join(base_path,file) for file in dirs]
    return full_paths
def get_daka_names_standard(path,nianji=None):
    wb = xlrd.open_workbook(path)
    sheet = wb.sheet_by_name("sheet")

    list_2019 = []
    list_2020 = []
    list_2021 = []
    list_phd = []

    #找到对应列
    id_col_num = 0   #学号列
    name_col_num = 1 #姓名列
    type_col_num = 3 #学生类别列
    for i,first_row_data in enumerate(sheet.row_values(0)):
        if  first_row_data == '学号':
            id_col_num = i
        elif first_row_data == '姓名':
            name_col_num = i
        elif first_row_data == '学生类别':
            type_col_num = i

    for i,(type_value,id_value) in enumerate(zip(sheet.col_values(type_col_num),sheet.col_values(id_col_num))):
        if str(type_value).find('博士') >= 0 or str(type_value).find('直博') >= 0:
            list_phd.append(sheet.cell_value(i,name_col_num))
        elif str(id_value)[:4] == '2019':
            list_2019.append(sheet.row_values(i,0,4))
            # list_2019.append(sheet.cell_value(i,name_col_num))

        elif str(id_value)[:4] == '2020':
            list_2020.append(sheet.cell_value(i,name_col_num))
        elif str(id_value)[:4] == '2021':
            list_2021.append(sheet.cell_value(i,name_col_num))
        # else:
        #     print("奇怪的值!")
        #     print(sheet.cell_value(i,id_col_num))
    list_dict = {'2019':list_2019,'2020':list_2020,'2021':list_2021,'博士':list_phd}
    return list_2019
    # print(list_2019)
    # exit()
    # if nianji:
    #     names = list_dict[nianji]
    #     names = remove_names(names, nianji)
    #     return names
    # else:
    #     nianjis = list_dict.keys()
    #     for nianji in nianjis:
    #         names = list_dict[nianji]
    #         names = remove_names(names, nianji)
    #         list_dict[nianji] = names
    # return list_dict

full_paths = get_excel_path(r"C:\Users\wmj\Desktop\excel")
contents = []
for path in full_paths:
    contents.extend(get_daka_names_standard(path))

contents = remove_dup(contents)
write_excel(contents,r"C:\Users\wmj\Desktop\results.xlsx")