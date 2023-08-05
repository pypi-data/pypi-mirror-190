import json
import shutil
import openpyxl  # excel 不能读取xls
import re
import os

BIG_TABLE_TITLE = ["编号", "名称及规格", "材料", "标准型号", "数量", "单位", "备注"]
SMALL_TABLE_TITLE = ["编号", "长度", "外径"]


class SmallTable:
    title = ["编号", "长度", "外径"]

    def __init__(self):
        self.items = []

    def add_item(self, id, length, diameter):
        item = {"编号": id, "长度": length, "外径": diameter}
        self.items.append(item)


class BigTable:
    def __init__(self):
        self.title = ["编号", "名称及规格", "材料", "标准型号", "数量", "单位", "备注"]
        self.items = []

    def add_item(self, id, name_size, material, standard_size, number, unit, remark):
        item = {"编号": id, "名称及规格": name_size, "材料": material, "标准型号": standard_size, "数量": number, "单位": unit,
                "备注": remark}
        self.items.append(item)


class AutocadImage:
    def __init__(self):
        self.big_table = None
        self.small_table = None
        self.install_image_id = ''
        self.self_image_id = ''


class CacheAutocadImage:
    def __init__(self):
        self.install_image_id_cell = None
        self.self_image_id_cell = None
        self.big_table_cells = {
            "编号": None, "名称及规格": None, "材料": None, "标准型号": None, "数量": None, "单位": None, "备注": None
        }
        self.small_table_cells = {
            "编号": [], "长度": [], "外径": []
        }


class AutocadExcelTool:
    @classmethod
    def get_install_image_id(self, sheet, str_cell):  # 获取安装图图号
        pass

    def get_self_image_id(self, sheet, str_cell):  # 获取本图图号
        pass


def is_big_table_cell_value_none(sheet, cell):
    if cell.column + len(BIG_TABLE_TITLE) >= sheet.max_column:
        return True

    for col in range(cell.column, cell.column + len(BIG_TABLE_TITLE)):
        if sheet[cell.row][col].value is None:
            return True

    return False


##如果只是统计数量
def calculate_big_tables(path):
    dirname = os.path.dirname(path)
    new_path = os.path.join(dirname, f"new_{os.path.basename(path)}")
    shutil.copy(path, new_path)

    wb = openpyxl.load_workbook(new_path)
    sheet = wb['Sheet1']
    # small_table = SmallTable()
    items = []

    for i in range(1, sheet.max_row + 1):
        for j in range(0, sheet.max_column):
            cell = sheet[i][j]
            if cell.value is None:
                continue

            match_result = re.match(r'^\\pxqc;\d+', cell.value, flags=0)
            if match_result is not None and not is_big_table_cell_value_none:

                is_name_size = False if re.match(r'^\\pxql;*', sheet[i][j + 1].value, flags=0) is None else True
                is_material = False if re.match(r'^\\pxqc;*', sheet[i][j + 2].value, flags=0) is None else True
                is_standard_size = False if re.match(r'^\\pxqc;*', sheet[i][j + 3].value, flags=0) is None else True
                is_number = False if re.match(r'^\\pxqc;\d+', sheet[i][j + 4].value, flags=0) is None else True
                is_unit = False if re.match(r'^\\pxqc;*', sheet[i][j + 5].value, flags=0) is None else True
                is_remark = False if re.match(r'^\\pxqc;*', sheet[i][j + 6].value, flags=0) is None else True

                if is_name_size and is_material and is_standard_size and is_number and is_unit and is_remark:
                    # small_table.add_item(sheet[i][j].value, sheet[i][j + 1].value, sheet[i][j + 2].value)
                    sub_items = []
                    for k in range(0, len(BIG_TABLE_TITLE)):
                        sub_items.append(sheet[i][j + k].value)
                        sheet[i][j + k].value = None
                    sub_items = tuple(sub_items)
                    items.append(sub_items)

    print(len(items))
    with open(os.path.join(os.path.dirname(path), "big_table.json"), 'w') as f:
        json.dump(items, fp=f, indent=1, ensure_ascii=False)

    wb.save(new_path)


##如果只是统计数量
def calculate_small_tables(path):
    dirname = os.path.dirname(path)
    new_path = os.path.join(dirname, f"new_{os.path.basename(path)}")
    shutil.copy(path, new_path)

    wb = openpyxl.load_workbook(new_path)
    sheet = wb['Sheet1']
    # small_table = SmallTable()
    items = []

    for i in range(1, sheet.max_row + 1):
        for j in range(0, sheet.max_column):
            cell = sheet[i][j]
            if cell.value is None:
                continue
            match_result = re.match(r'<\d+>', cell.value, flags=0)
            if match_result is not None:
                if j + 2 < sheet.max_column and str(sheet[i][j + 1].value).isdigit() and str(
                        sheet[i][j + 2].value).isdigit():
                    # small_table.add_item(sheet[i][j].value, sheet[i][j + 1].value, sheet[i][j + 2].value)
                    items.append((sheet[i][j].value, sheet[i][j + 1].value, sheet[i][j + 2].value))
                    print(sheet[i][j].value, sheet[i][j + 1].value, sheet[i][j + 2].value)
                    sheet[i][j].value = None
                    sheet[i][j + 1].value = None
                    sheet[i][j + 2].value = None

    print(len(items))
    with open(os.path.join(os.path.dirname(path), "small_table.json"), 'w') as f:
        json.dump(items, fp=f, indent=1, ensure_ascii=False)

    wb.save(new_path)


# def get_excel(path):
#     wb = openpyxl.load_workbook(path)
#     sheet = wb['Sheet1']
#     print(f"max row:{sheet.max_row} min row:{sheet.min_row}")
#     print(f"max col:{sheet.max_column} min col:{sheet.min_column}")
#
#     auto_images = []
#     error_infos = []
#     last_cai = None
#
#     ai = AutocadImage()
#     cai = CacheAutocadImage()
#     big_table = BigTable()
#     small_table = SmallTable()
#
#     for i in range(1, sheet.max_row + 1):
#         for j in range(0, sheet.max_column):
#             a = sheet[i][j]
#             if a.value == "编号":
#                 flag = 0 # 0初始 1big 2small
#                 index = 1
#                 max_index = 7
#                 for k in range(j+1,sheet.max_column):
#                     if flag==1 or (flag == 0 and sheet[i][k].value == BIG_TABLE_TITLE[index]):
#                         flag = 1
#                         cai.big_table_cells["编号"] = a
#                         cai.big_table_cells[BIG_TABLE_TITLE[index]] = sheet[i][k]
#                         index += 1
#                     elif flag==0 and sheet[i][k].value == SMALL_TABLE_TITLE[index]:
#                         flag = 2
#                         cai.small_table_cells["编号"].append(a)
#                         cai.small_table_cells[BIG_TABLE_TITLE[index]].append(sheet[i][k])
#                         index += 1
#


# 开始进入检索模式

# for row in range(1,sheet)

# for col in sheet.columns:
#     for row in sheet.ro

def main():
    path = r'C:\Users\wmj\Desktop\sheettest.xlsx'
    small_table_path = r"C:\Users\wmj\Desktop\small_table.xlsx"
    big_table_path = r"C:\Users\wmj\Desktop\big_tables.xlsx"
    calculate_small_tables(big_table_path)


main()
