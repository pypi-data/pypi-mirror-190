import json
import os
import re
import _thread
from autofix_excel_column_size import CXlAutofit

import openpyxl  # excel 不能读取xls
import numpy as np
from pyautocad import Autocad, APoint

BIG_TABLE_TITLE = ["编号", "名称及规格", "材料", "标准型号", "数量", "单位", "备注"]
SMALL_TABLE_TITLE = ["编号", "长度", "外径"]
INSTALL_TITLE = ["行", "列", "安装图图号", "本图图号"]


class Tool:
    def __init__(self,root=""):
        self.acad = Autocad(create_if_not_exists=True)
        self.acad.prompt("Hello, Autocad from Python\n")
        print(self.acad.doc.Name)

        self.original_point = (0.0, 0.0)
        self.width = 470  # x
        self.height = 347  # y

        self.small_save_json_path = os.path.join(root,r"small_table_py.json")
        self.small_save_excel_path = os.path.join(root,r"small_table_all.xlsx")

        self.big_save_json_path = os.path.join(root,r"big_table_py.json")
        self.big_save_excel_path = os.path.join(root,r"big_table_all.xlsx")

        self.install_save_json_path = os.path.join(root,r"install_table_py.json")
        self.install_save_excel_path = os.path.join(root,r"install_all.xlsx")

    def calculate_r_c(self, ordinate):
        x, y = ordinate[0], ordinate[1]
        r = int((y - self.original_point[1]) / self.height) + 1
        c = int((x - self.original_point[0]) / self.width) + 1
        return str(r), str(c)

    def show_small_dic(self):
        all_dic = self.get_all_dic('all')
        print(f"-------------------------------"
              f"Test InsertionPoint and TextAlignmentPoint"
              f"-------------------------------")
        for row in all_dic.keys():
            for col in all_dic[row].keys():
                if all_dic[row][col][1][0] <= 0 or all_dic[row][col][1][1] <= 0:
                    print(f"InsertionPoint has 0: row:{row},col:{col},insertionpoint:{all_dic[row][col][1]}")
                if all_dic[row][col][2][0] <= 0 or all_dic[row][col][2][1] <= 0:
                    print(f"TextAlignmentPoint has 0: row:{row},col:{col},TextAlignmentPoint:{all_dic[row][col][1]}")

    def get_dic_save_contents(self, item, dic_saved_type):
        if dic_saved_type == "TextString":
            return item.TextString
        else:
            return {'text': item.TextString, 'x': item.InsertionPoint[0], 'y': item.InsertionPoint[1]}

    def get_all_install_dic(self):
        if os.path.exists(self.install_save_json_path):
            with open(self.install_save_json_path, encoding='utf-8') as f:
                install_num = json.loads(f.read())
            return install_num

        install_num = {}
        for item in self.acad.iter_objects():
            if "-" in item.TextString:
                r, c = self.calculate_r_c(item.InsertionPoint)
                # 创建字典
                if r not in install_num.keys():
                    install_num[r] = {}
                    install_num[r][c] = []
                else:
                    if c not in install_num[r].keys():
                        install_num[r][c] = []

                install_num[r][c].append(item.TextString)

        with open(self.install_save_json_path, 'w', encoding='utf-8') as f:
            json.dump(install_num, ensure_ascii=False, indent=1, fp=f)

        return install_num

    def save_install_num(self):
        excel_rows = []
        install_num = self.get_all_install_dic()

        # 查看
        for row in install_num.keys():
            for col in install_num[row].keys():
                nums = install_num[row][col]
                if len(nums) != 2:
                    print("不等于2")
                else:
                    i1 = nums[1] if len(nums[0]) > len(nums[1]) else nums[0]
                    i2 = nums[0] if len(nums[0]) > len(nums[1]) else nums[1]
                    excel_rows.append([row, col, i1, i2])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(INSTALL_TITLE)
        for excel_row in excel_rows:
            ws.append(excel_row)
        wb.save(self.install_save_excel_path)

    def combine_install_num(self, excel_type='big'):
        if excel_type == 'big':
            origin_path = self.big_save_excel_path
        else:
            origin_path = self.small_save_excel_path

        wb = openpyxl.load_workbook(origin_path)
        ws = wb.active

        all_install_dic = self.get_all_install_dic()

        ws.insert_cols(idx=3)
        ws.insert_cols(idx=3)
        ws[1][2].value, ws[1][3].value = "安装图图号", "本图图号"
        # ws.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        for i, row in enumerate(ws.rows):
            r = row[0].value
            c = row[1].value

            if i == 0:
                continue

            install_ids = all_install_dic[r][c]
            id1 = install_ids[0]
            id2 = install_ids[1]

            row[2].value = id1
            row[3].value = id2

        new_path = os.path.join(os.path.dirname(origin_path), 'new_' + os.path.basename(origin_path))
        wb.save(new_path)

        if excel_type == 'big':
            self.process_sheet_style_center(new_path, 'Sheet')
            self.process_sheet_style_center(new_path, 'big_unparsed')
            self.process_sheet_style_data_type(new_path, 'Sheet', [1, 2, 5], excel_type=excel_type)
        else:
            self.process_sheet_style_center(new_path, 'Sheet')
            self.process_sheet_style_data_type(new_path, 'Sheet', [1, 2, 5, 6, 7], excel_type=excel_type)

    def process_sheet_style_install(self):
        # new_path = os.path.join(os.path.dirname(self.install_save_excel_path),
        #                         'new_' + os.path.basename(self.install_save_excel_path))
        self.process_sheet_style_center(self.install_save_excel_path, "Sheet")
        self.process_sheet_style_data_type(self.install_save_excel_path, "Sheet", [1, 2])

    def process_sheet_style_center(self, excel_path, sheet_name):
        Entity = CXlAutofit()
        Entity.style_excel(excel_path, sheet_name)

    def process_sheet_style_data_type(self, excel_path, sheet_name, columns, excel_type='big'):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name]

        for i in range(2, ws.max_row + 1):
            # 遍历每一列的全部行
            for j in range(1, ws.max_column + 1):
                if j in columns and ws.cell(row=i, column=j).value is not None:
                    value = ws.cell(row=i, column=j).value
                    if value[0] == '<' and value[-1] == '>':
                        value = value[1:-1]
                    ws.cell(row=i, column=j).value = int(value)
        wb.save(excel_path)

    def process_small(self):
        self.process_sheet_style_center(self.small_save_excel_path, "sheet")
        self.process_sheet_style_data_type(self.small_save_excel_path)

    def combine_all_sheets_to_one_excel(self):  # 把newcombinae

        small_path = os.path.join(os.path.dirname(self.small_save_excel_path),
                                  'new_' + os.path.basename(self.small_save_excel_path))
        big_path = os.path.join(os.path.dirname(self.big_save_excel_path),
                                'new_' + os.path.basename(self.big_save_excel_path))
        install_path = self.install_save_excel_path
        combine_excel_path = os.path.join(os.path.dirname(self.big_save_excel_path),
                                          'final.xlsx')

        from excel_files_combine import CombineExcelFile
        cef = CombineExcelFile()

        cef.combine_by_paths(paths=[install_path, small_path], sheet_names=['install', 'small'],
                             combine_excel_path=combine_excel_path)

    def get_all_dic(self, type='small', dic_saved_type='TextString'):
        if type == 'small':
            path = self.small_save_json_path
        else:
            path = self.big_save_json_path
        if os.path.exists(path):
            with open(path, encoding='utf-8') as f:
                all_dic = json.loads(f.read())
                return all_dic

        all_dic = {}  # r c
        for item in self.acad.iter_objects():
            r, c = self.calculate_r_c(item.InsertionPoint)

            # 创建字典
            if r not in all_dic.keys():
                all_dic[r] = {}
                all_dic[r][c] = []
            else:
                if c not in all_dic[r].keys():
                    all_dic[r][c] = []

            # print(f"r:{r},c:{c},str:{item.TextString}")
            all_dic[r][c].append(self.get_dic_save_contents(item, dic_saved_type))

        return all_dic

    def save_small_table(self):
        print(f"Start get all dict")
        all_dic = self.get_all_dic(type='small')
        with open(self.small_save_json_path, "w") as f:
            json.dump(all_dic, fp=f, indent=1, ensure_ascii=False)
        print(f"All dict has saved")

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = "行"
        ws['B1'] = "列"
        ws['C1'] = "编号"
        ws['D1'] = "长度"
        ws['E1'] = "外径"
        title_length = 3

        print(f"Start generate small table excel")

        for row in all_dic.keys():
            for col in all_dic[row].keys():
                sub_items = all_dic[row][col]
                if len(sub_items) % title_length != 0:
                    print(f'Wrong small images: row:{row} col:{col} num:{len(sub_items)}')
                    continue
                for i in range(0, len(sub_items), title_length):
                    if sub_items[i] in SMALL_TABLE_TITLE:
                        continue
                    ws.append([row, col, sub_items[i], sub_items[i + 1], sub_items[i + 2]])
        wb.save(self.small_save_excel_path)

        print(f"Small table excel has saved in {self.small_save_json_path}")

    def _put_proper_col_on_row(self, rows, need_to_restruct_rows, names):
        for i in need_to_restruct_rows:
            row = rows[i]
            print(row)

            new_row = []
            for row in rows:
                for i in row:
                    if not isinstance(i, str):
                        new_row.append(i)
            row = new_row

            # row = sorted(row, key=lambda i: i[1][0])  # 按照位置排
            row.sort(key=lambda x: x[1][0])

            if len(row) > len(names):  # 超过了
                new_row = []
                for row_item in row:
                    if row[0] != "":
                        new_row.append(row_item)
                row = new_row

            if len(row) == len(names):  # 是一样的长度，那就OK了
                rows[i] = row
                continue

            else:  # 不一样的长度要定位
                print(row)
                new_row = [''] * len(names)
                for r in row:
                    min_distance_pos = np.argmin(np.array([abs(r[1][0] - name[1][0]) for name in names]))
                    if new_row[min_distance_pos] != '':
                        if new_row[min_distance_pos][0] == '\\pxqc;':
                            new_row[min_distance_pos][0] = r
                            print("剔除占位符")
                        else:
                            if min_distance_pos + 1 >= len(new_row):
                                print("hhhhh")
                            new_row[min_distance_pos + 1] = r
                    else:
                        new_row[min_distance_pos] = r
                rows[i] = new_row

        return rows

    def _append_big_correct_row(self, rows, items, title_length=7):
        need_to_restruct_rows = []
        if len(items) == 0:
            for index, row in enumerate(rows):
                if len(row) == title_length:
                    continue
                need_to_restruct_rows.append(index)
            return rows, need_to_restruct_rows

        for item in items:
            is_match = False
            for index, row in enumerate(rows):
                need_to_restruct_rows.append(index)
                if item[1][1] == row[0][1][1]:
                    row.append(item)
                    is_match = True
                    break
            if not is_match:
                print("something wrong without match")
        return rows, need_to_restruct_rows

    def seat_princple(self, name, text):  # 返回False表示不匹配 返回True表示可以匹配
        # ["编号", "名称及规格", "材料", "标准型号", "数量", "单位", "备注"]
        if name == '编号':
            flag = re.match(r'\\pxqc;\d+', text)
            return flag is not None
        elif name == '名称及规格':
            if '单头丝短接 PN16 G1' in text:
                flag = True
            else:
                flag = '\pxql;' in text
            return flag
        elif name == '材料':
            return True
        elif name == '标准型号':
            text = text.replace(r"\pxqc;", '')
            flag = not text.isdigit()
            return flag
        elif name == '数量':
            text = text.replace(r"\pxqc;", '')
            flag = text.isdigit()
            return flag
        elif name == '单位':
            flag = text == "\\pxqc;个" or text == "\\pxqc;米"
            text = text.replace(r"\pxqc;", '')
            if not flag and len(text) < 2:
                print(text)
            return flag
        elif name == '备注':
            return True

    def seat_pos(self, items, names):
        rest_items_len = len(items)
        rest_names_len = len(names)
        excel_row = [''] * 7
        pos = -1

        for i in range(len(items)):
            item = items[i]
            item_names = names[(pos + 1):(pos + 1 + rest_names_len - rest_items_len + 1)]
            relative_pos = self.seat_one_pos(item, item_names)
            if relative_pos == -1:
                print("难以匹配，退出")
                return None
            pos = len(names) - rest_names_len + relative_pos
            excel_row[pos] = item['text']

            rest_names_len = len(names) - pos - 1
            rest_items_len -= 1

        return excel_row

    def seat_one_pos(self, item, names):
        for i, name in enumerate(names):
            if self.seat_princple(name, item['text']):
                return i
        return -1

    def generate_excel_rows(self, sorted_dic):
        excel_rows = []
        special_canot_exe = []
        for row in sorted_dic.keys():
            for col in sorted_dic[row].keys():
                for y, one_line_items in sorted_dic[row][col].items():
                    if len(one_line_items) == 7:
                        excel_row = [item['text'] for item in one_line_items]
                        res = [row, col]
                        res.extend(excel_row)
                        excel_rows.append(res)
                    elif len(one_line_items) < 7:
                        excel_row = self.seat_pos(one_line_items, BIG_TABLE_TITLE)
                        if excel_row is not None:
                            res = [row, col]
                            res.extend(excel_row)
                            excel_rows.append(res)
                        else:
                            res = [row, col]
                            res.extend([item['text'] for item in one_line_items])
                            special_canot_exe.append(res)
                            # special_canot_exe.append({
                            #     'r': row,
                            #     'c': col,
                            #     'items': one_line_items
                            # })
                    else:
                        res = [row, col]
                        res.extend([item['text'] for item in one_line_items])
                        special_canot_exe.append(res)
        print(len(excel_rows))
        print(len(special_canot_exe))
        print(special_canot_exe)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws2 = wb.create_sheet('big_unparsed', 2)

        self._add_title(ws, BIG_TABLE_TITLE)
        self._add_title(ws2, BIG_TABLE_TITLE)

        excel_rows = self._replace_all_unused_text(excel_rows)
        special_canot_exe = self._replace_all_unused_text(special_canot_exe)

        for excel_row in excel_rows:
            ws.append(excel_row)

        for excel_row in special_canot_exe:
            ws2.append(excel_row)

        wb.save(self.big_save_excel_path)

    def _add_title(self, ws, names):
        ws['A1'] = "行"
        ws['B1'] = "列"
        ws['C1'] = names[0]
        ws['D1'] = names[1]
        ws['E1'] = names[2]
        ws['F1'] = names[3]
        ws['G1'] = names[4]
        ws['H1'] = names[5]
        ws['I1'] = names[6]

    def _replace_all_unused_text(self, excel_rows):
        new_excel_rows = []
        for excel_row in excel_rows:
            new_excel_row = []
            for i in excel_row:
                i = i.replace(r"\pxqc;", "")
                i = i.replace(r"\pxql;", '')
                i = i.replace(r"{\W0.48;", '')
                i = i.replace(r"{\W0.79;", '')
                i = i.replace(r"{\W0.61;", '')
                i = i.replace(r"{\W0.78;", '')
                i = i.replace(r"\fISOCPEUR|b0|i0|c134|p34;%%C\fSimSun|b0|i0|c134|p2;", "φ")
                i = i.replace(r"\A1;", "")
                i = i.replace(r"\P", "")
                i = i.replace(r"{\W0.78;", "")
                i = i.replace(r"{\H0.7x;\S1/2;}", "1/2")
                if len(i) > 0 and i[-1] == '}':
                    i = i[:-1]
                new_excel_row.append(i)
            new_excel_rows.append(new_excel_row)
        return new_excel_rows

    def show_sorted_dic(self, sorted_dic):
        for row in sorted_dic.keys():
            for col in sorted_dic[row].keys():
                for y, one_line_items in sorted_dic[row][col].items():
                    if len(one_line_items) > 7:
                        print('Over 7')
                        print(one_line_items)
                    elif len(one_line_items) < 2:
                        print(len(one_line_items))
                        # print('Less 7')

    def remove_null_over_less_title_length(self, sorted_dic, title_length=7):
        remoted_dic = {}

        for row in sorted_dic.keys():
            remoted_dic[row] = {}
            for col in sorted_dic[row].keys():
                remoted_dic[row][col] = {}
                for y, one_line_items in sorted_dic[row][col].items():
                    if len(one_line_items) != title_length:  # 超过或者小于

                        new_one_line_items = []
                        for item in one_line_items:
                            if item['text'] != '\\pxqc;':
                                new_one_line_items.append(item)

                        if len(new_one_line_items) > 0:
                            remoted_dic[row][col][y] = new_one_line_items
                    else:
                        remoted_dic[row][col][y] = sorted_dic[row][col][y]

        return remoted_dic

    def delete_duplicate(self, data):
        func = lambda x, y: x + [y] if y not in x else x
        from functools import reduce
        data = reduce(func, [[], ] + data)
        return data

    def sort_all_same_row_dic(self,
                              all_same_row_dic):  # 构 all_same_row_dic[3][2][345.678] : [{'x':12,'y':345.678,'text':'编号'},...]
        # 用x排序 去掉重复的x
        for row in all_same_row_dic.keys():
            for col in all_same_row_dic[row].keys():
                for y, one_line_items in all_same_row_dic[row][col].items():
                    sorted_one_line_items = sorted(one_line_items, key=lambda i: i['x'])
                    # 去重

                    without_dup_line_items = self.delete_duplicate(sorted_one_line_items)
                    all_same_row_dic[row][col][y] = without_dup_line_items
        return all_same_row_dic

    def combine_save_y(self, all_dic):
        # 去除掉 编号这些
        all_same_row_dic = {}
        for row in all_dic.keys():
            all_same_row_dic[row] = {}
            for col in all_dic[row].keys():
                all_same_row_dic[row][col] = {}

                # 构 all_same_row_dic[3][2][345.678] : [{'x':12,'y':345.678,'text':'编号'},...]
                image_items = all_dic[row][col]
                for item in image_items:
                    # 把编号这些忽略掉
                    if item['text'] in BIG_TABLE_TITLE:
                        continue

                    y = item['y']
                    if y in all_same_row_dic[row][col].keys():
                        all_same_row_dic[row][col][y].append(item)
                    else:
                        all_same_row_dic[row][col][y] = [item]

        return all_same_row_dic

    def order_x_for_all_same_row_dic(self, all_same_row_dic):
        pass

    def exeu_wrong_big_dic1(self):
        with open(self.big_save_json_path) as f:
            all_dic = json.loads(f.read())
        with open(self.big_wrong_big_dic_path) as f:
            wrong_dic = json.loads(f.read())

        title_length = 7

        wb = openpyxl.Workbook()
        ws = wb.active

        for (r, c) in wrong_dic:
            sub_items = all_dic[r][c]
            print(len(sub_items))
            rows = []

            i = 0
            id_no = 1
            names = []
            remains_items = []
            while (i < len(sub_items)):

                # 如果是编号...
                if sub_items[i][0] in BIG_TABLE_TITLE:
                    names = [sub_items[i] for i in range(i, i + 7)]
                    i = i + title_length
                    continue

                row = [''] * title_length
                # 正常数量
                if i + title_length < len(sub_items) and sub_items[i][0] == f'\\pxqc;{id_no}' and \
                        sub_items[i + title_length][0] == f'\\pxqc;{id_no + 1}':

                    row[0] = sub_items[i]  # 标号的
                    for j in range(i + 1, i + title_length):
                        # y不吻合剔除
                        if sub_items[j][1][1] != row[0][1][1]:
                            remains_items.append(sub_items[j][1][1])
                            print("The 7 num true item is wrong")
                        else:
                            row[j - i] = sub_items[j]

                    rows.append(row)
                    i = i + title_length
                    id_no += 1
                    continue
                elif sub_items[i][0] == f'\\pxqc;{id_no}':
                    is_find_next_id_not = False
                    for j in range(i + 1, min(i + title_length, len(sub_items))):
                        if sub_items[j][0] == f'\\pxqc;{id_no + 1}':
                            # 根据names的坐标放在合适位置,先暂时放在该行

                            new_row = []
                            new_row.append(sub_items[i])
                            for row_item in sub_items[i:j]:
                                if row_item[1][1] != new_row[0][1][1]:
                                    remains_items.append(row_item)
                                    print("The Second item is wrong")
                                else:
                                    new_row.append(sub_items[j])
                            row = new_row

                            rows.append(row)
                            i = j
                            id_no += 1
                            # 回到大循环
                            is_find_next_id_not = True
                            break
                    if not is_find_next_id_not:
                        row = sub_items[i:min(i + title_length, len(sub_items))]
                        rows.append(row)
                        print(f"do not find the next id_no:{id_no + 1}")
                        i = min(i + title_length, len(sub_items))
                    continue
                else:
                    print("有剩余位置的item")
                    remains_items.append(sub_items[i:])
                    # remain
                break

            rows, need_to_restruct_rows = self._append_big_correct_row(rows, remains_items)
            rows = self._put_proper_col_on_row(rows, need_to_restruct_rows, names)

            for row in rows:
                res = [r, c]

                for i in row:
                    if isinstance(i, str):
                        res.append(i)
                    else:
                        res.append(i[0])
                ws.append(row)

        wb.save(r'C:\Users\wmj\Desktop\dad1\hhh.xlsx')

    def save_big_table(self):
        print(f"Start get all dict")
        all_dic = self.get_all_dic(type='big', dic_saved_type='InsertionPoint')
        with open(self.big_save_json_path, "w") as f:
            json.dump(all_dic, fp=f, indent=1, ensure_ascii=False)
        print(f"All dict has saved")

        wb = openpyxl.Workbook()
        ws = wb.active

        wrong_big_dic = []

        ws['A1'] = "Row"
        ws['B1'] = "Col"
        ws['C1'] = "编号"
        ws['D1'] = "名称及规格"
        ws['E1'] = "材料"
        ws['F1'] = "标准型号"
        ws['G1'] = "数量"
        ws['H1'] = "单位"
        ws['I1'] = "备注"

        title_length = 7

        print(f"Start generate big table excel")

        for row in all_dic.keys():
            for col in all_dic[row].keys():
                sub_items = all_dic[row][col]
                if len(sub_items) % title_length != 0:
                    # 长度不同，单独处理
                    wrong_big_dic.append((row, col))
                    print(f'Wrong big images: row:{row} col:{col} num:{len(sub_items)}')
                    continue

                for i in range(0, len(sub_items), title_length):
                    if sub_items[i][0] in SMALL_TABLE_TITLE:
                        continue

                    res = [row, col]
                    res.extend([sub_items[j][0] for j in range(i, i + title_length)])
                    ws.append(res)
        wb.save(self.big_save_excel_path)
        with open(self.big_wrong_big_dic_path, 'w') as f:
            json.dump(wrong_big_dic, fp=f, indent=1, ensure_ascii=False)

        print(f"Big table excel has saved in {self.big_save_excel_path}")

    #
    # def save_big_table1(self):
    #     save_json_path = r"C:\Users\wmj\Desktop\dad\big_table_py.json"
    #     save_excel_path = r"C:\Users\wmj\Desktop\dad\big_table_all.xlsx"
    #
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #
    #     ws.title = "small_table"
    #     ws['A1'] = "Row"
    #     ws['B1'] = "Col"
    #     ws['C1'] = "编号"
    #     ws['D1'] = "名称及规格"
    #     ws['E1'] = "材料"
    #     ws['F1'] = "标准型号"
    #     ws['G1'] = "数量"
    #     ws['H1'] = "单位"
    #     ws['I1'] = "备注"
    #     all_dic = {}  # r c
    #
    #     if not os.path.exists(save_json_path):
    #         for item in self.acad.iter_objects():
    #             r, c = self.calculate_r_c(item.InsertionPoint)
    #
    #             # 创建字典
    #             if r not in all_dic.keys():
    #                 all_dic[r] = {}
    #                 all_dic[r][c] = []
    #             else:
    #                 if c not in all_dic[r].keys():
    #                     all_dic[r][c] = []
    #
    #             print(f"r:{r},c:{c},str:{item.TextString}")
    #
    #             all_dic[r][c].append((item.TextString, item.InsertionPoint[0], item.InsertionPoint[1]))
    #
    #             with open(save_json_path, "w") as f:
    #                 json.dump(all_dic, fp=f, indent=1, ensure_ascii=False)
    #     else:
    #         with open(save_json_path) as f:
    #             all_dic = json.loads(f.read())
    #
    #     for row in all_dic.keys():
    #         for col in all_dic[row].keys():
    #             sub_items = all_dic[row][col]
    #
    #             names = []
    #             index = -1
    #             for i, si in enumerate(sub_items):
    #                 if si[0] in BIG_TABLE_TITLE:
    #                     names.append(si)
    #                     index = i + 1
    #
    #                 if len(names) == 7:
    #                     break
    #
    #             for i in range(index + 6, len(sub_items), 7):
    #                 # i = index+6
    #                 if re.match(r'\\pxqc;\d+', sub_items[i - 6][0]):
    #                     id = sub_items[i - 6][0]
    #                 else:
    #                     id = ''
    #                     print("something wrong")
    #
    #                 name_size = sub_items[i - 5][0]
    #                 mat = sub_items[i - 4][0]
    #                 stand = sub_items[i - 3][0]
    #                 num = sub_items[i - 2][0]
    #                 unit = sub_items[i - 1][0]
    #                 remark = sub_items[i][0]
    #                 ws.append([row, col, id, name_size, mat, stand, num, unit, remark])
    #                 index = i
    #                 #
    #
    #             remain_res = [row, col]
    #             for i in range(index + 1, len(sub_items)):
    #                 remain_res.append(sub_items[i][0])
    #
    #             ws.append(remain_res)
    #
    #     wb.save(save_excel_path)
    #
    # def save_big_table(self):
    #     save_json_path = r"C:\Users\wmj\Desktop\dad\big_table_py.json"
    #     save_excel_path = r"C:\Users\wmj\Desktop\dad\big_table_all.xlsx"
    #
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #
    #     ws.title = "small_table"
    #     ws['A1'] = "Row"
    #     ws['B1'] = "Col"
    #     ws['C1'] = "编号"
    #     ws['D1'] = "名称及规格"
    #     ws['E1'] = "材料"
    #     ws['F1'] = "标准型号"
    #     ws['G1'] = "数量"
    #     ws['H1'] = "单位"
    #     ws['I1'] = "备注"
    #     all_dic = {}  # r c
    #
    #     if not os.path.exists(save_json_path):
    #         for item in self.acad.iter_objects():
    #             r, c = self.calculate_r_c(item.InsertionPoint)
    #
    #             # 创建字典
    #             if r not in all_dic.keys():
    #                 all_dic[r] = {}
    #                 all_dic[r][c] = []
    #             else:
    #                 if c not in all_dic[r].keys():
    #                     all_dic[r][c] = []
    #
    #             print(f"r:{r},c:{c},str:{item.TextString}")
    #
    #             all_dic[r][c].append((item.TextString, item.InsertionPoint[0], item.InsertionPoint[1]))
    #
    #             with open(save_json_path, "w") as f:
    #                 json.dump(all_dic, fp=f, indent=1, ensure_ascii=False)
    #     else:
    #         with open(save_json_path) as f:
    #             all_dic = json.loads(f.read())
    #
    #     for row in all_dic.keys():
    #         for col in all_dic[row].keys():
    #             sub_items = all_dic[row][col]
    #
    #             names = []
    #             index = -1
    #             no_si = None
    #             for i, si in enumerate(sub_items):
    #                 if si[0] in BIG_TABLE_TITLE:
    #                     names.append(si)
    #                     index = i + 1
    #                     if si[0] == "编号":
    #                         no_si = si
    #
    #                 if len(names) == 7:
    #                     break
    #
    #             last_position = index
    #             id_final_positions = []
    #             for id_no in range(1, 20):
    #                 id_positions = []
    #                 for i in range(last_position + 1, len(sub_items)):
    #                     if str(sub_items[i][0]) == f'\\pxqc;{id_no}':
    #                         id_positions.append(i)
    #
    #                 if len(id_positions) == 0:
    #                     break
    #                 elif len(id_positions) == 1:
    #                     id_final_positions.extend(id_positions)
    #                 else:
    #                     id_position = self._choose_id_no(sub_items, id_positions, no_si)
    #                     id_final_positions.append(id_position)
    #
    #             rows = []
    #
    #             pos2 = -1
    #             for i in range(1, len(id_final_positions)):
    #                 pos2 = id_final_positions[i]
    #                 pos1 = id_final_positions[i - 1]
    #                 if pos2 - pos1 == 7:
    #                     row_cells = [row, col].extend([sub_items[index][0] for index in range(pos1, pos2)])
    #                 else:
    #                     a = self._get_proper_row_cells(sub_items, names, pos1, pos2)
    #                     row_cells = [row, col].extend(a)
    #                 rows.append(row_cells)
    #
    #             pos1 = pos2
    #             pos2 = len(id_final_positions)
    #             if pos2 - pos1 == 7:
    #                 row_cells = [row, col].extend([sub_items[index][0] for index in range(pos1, pos2)])
    #             else:
    #                 row_cells = [row, col].extend(self._get_proper_row_cells(sub_items, names, pos1, pos2))
    #             rows.append(row_cells)
    #             print(rows)

    # # 找编号位置
    # id_ord = None
    # for si in sub_items:
    #     if si[0] == "编号":
    #         id_ord = si
    #         break
    #
    # # 找标题
    # names = []
    # for si in sub_items:
    #     if self._is_same_ord(si[1], id_ord[1]):
    #         names.append(si[1])
    # names_num = len(names)
    # names.sort()
    #
    # # 找有几个项
    # ids = []
    # for si in sub_items:
    #     if self._is_same_ord(si[2], id_ord[2]):
    #         ids.append(si[2])
    # ids_num = len(ids)
    # ids.sort()
    #
    # print(f"names num:{names_num}  ids num:{ids_num}")
    # res = [([''] * ids_num) for i in range(names_num)]
    #
    # # 构造res
    # for si in sub_items:
    #     row, col = self._get_ord_r_c(names, ids, si[1], si[2])
    #     res[row-1][col-1] = si[0]

    #         for i in res:
    #             ws.append(i)
    # wb.save(save_excel_path)

    def _is_same_ord(self, o1, o2):
        if abs(o1 - o2) < 1:
            return True
        else:
            return False

    def _get_ord_r_c(self, names, ids, x, y):
        r, c = -1, -1
        for i, id in enumerate(ids):
            if self._is_same_ord(id, x):
                c = i + 1
        for i, name in enumerate(names):
            if self._is_same_ord(name, y):
                r = i + 1
        return r, c

    def _choose_id_no(self, sub_items, id_positions, no_si):
        distances = []
        for p in id_positions:
            si = sub_items[p]
            distance = abs(no_si[1] - si[1])
            distances.append(distance)
        i = np.argmax(np.array(distances))
        return id_positions[i]

    def _calculate_distance(self, sub_items, pos, sis):
        distances = []
        for name_si in sis:
            si = sub_items[pos]
            distance = abs(name_si[1] - si[1])
            distances.append(distance)
        i = np.argmax(np.array(distances))
        return i

    def _get_proper_row_cells(self, sub_items, names, pos1, pos2):
        res = [''] * 7
        k = 0
        res[k] = sub_items[pos1][0]
        k += 1
        for i in range(pos1 + 1, pos2):
            if i == pos2 - 1:
                res[6] = sub_items[i][0]
                break
            pos = self._calculate_distance(sub_items, i, names[k:]) + k
            res[pos] = sub_items[i][0]

            # 后面剩下的
            if pos2 - i == 7 - pos:
                res = res[:pos + 1].extend([sub_items[index][0] for index in range(pos + 1, pos2)])
                break

        return res
