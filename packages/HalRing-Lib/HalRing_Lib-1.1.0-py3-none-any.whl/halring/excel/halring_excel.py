# -*- coding:UTF-8 -*-
"""
__title__ = '对Excel单元格的读写'
__author__ = 'wu.keke'
__mtime__ = '2020/12/2'
#
      1.pip install xlrd
      2.pip install loguru
      3.pip install openpyxl
"""
import datetime
import os
import traceback

import xlrd
from loguru import logger
from xlrd import xldate_as_tuple
from openpyxl.reader.excel import load_workbook


class excelUtil():
    """ excel 帮助类 """
    def __init__(self, xlsxPath, sheetname=None):
        """

        :param xlsxPath:
        :param sheetname: sheetname若为None 则默认读取索引为0即第一个sheet的内容
        """

        self._xlsxPath = xlsxPath
        if sheetName is not None:
            self._sheetname = sheetname

    def readExcel(self):
        """
        读取excel
        :return: 返回数据为二维字典 i,j分别为excel的行和列的索引 索引从0开始 data[i][j]=cell_value
        """
        rows_data = {}
        try:
            if self._xlsxPath is None:
                return logger.error("路径不能为None")
            else:
                if not os.path.exists(self._xlsxPath):
                    return logger.error(f"{self._xlsxPath} 路径不存在")

            excl = xlrd.open_workbook(self._xlsxPath)
            if self._sheetname is not None:
                table = excl.sheet_by_name(self._sheetname)
            else:
                table = excl.sheet_by_index(0)

            rowNum = table.nrows
            colNum = table.ncols


            for i in range(rowNum):
                sheet_data = {}

                for j in range(colNum):
                    c_type = table.cell(i, j).ctype
                    c_cell = table.cell_value(i, j)
                    if c_type == 2 and c_cell % 1 == 0:
                        c_cell = int(c_cell)
                    elif c_type == 3:
                        date = datetime.datetime(*xldate_as_tuple(c_cell, 0))
                        c_cell = date.strftime('%Y%d%m %H:%M%S')
                    elif c_type == 4:
                        c_cell = True if c_cell == 1 else False

                    sheet_data[j] = c_cell
                    rows_data[i] = sheet_data
        except Exception as ex:
            logger.error("\tError %s\n" % ex)
            logger.error(traceback.format_exc())

        return rows_data

    def write_cell(self, row, col, value):
        """
        写入excel
        :param row: 要修改的单元格的行号 从0开始
        :param col: 要修改的单元格列号 从0开始
        :param value: 要写进单元格的值
        :return:
        """
        try:
            if self._xlsxPath is None:
                return logger.error("路径不能为None")
            else:
                if not os.path.exists(self._xlsxPath):
                    return logger.error(f"{self._xlsxPath} 文件不存在")
            if self._sheetname is None:
                return logger.error("sheetName不能为None")

            excl = load_workbook(self._xlsxPath)
            sheet = excl[self._sheetname]
            sheet.cell(row + 1, col + 1, value)
            excl.save(self._xlsxPath)
            return "SUCCESS"

        except Exception as ex:
            if "[Errno 13] Permission denied" in str(ex):
                return logger.error(f"{self._xlsxPath} 文件被占用,请先关闭文件")
            else:
                logger.error("\tError %s\n" % ex)
                return logger.error(traceback.format_exc())


if __name__ == '__main__':
    xlsxPath = "D:\\2020\\NGAir方法整理.xlsx"
    sheetName = "utils"
    excel_util = excelUtil(xlsxPath, sheetName)
    # 结果二维字典
    # data = excel_util.readExcel()
    # print(data[1][1])
    # 修改以及添加单元格的值
    excel_util.write_cell(47, 1, "三翻四复")
    excel_util.write_cell(47, 2, "快接啊号地块计划")
    excel_util.write_cell(47, 3, "好人")
    excel_util.write_cell(48, 1, "搜索vs")
