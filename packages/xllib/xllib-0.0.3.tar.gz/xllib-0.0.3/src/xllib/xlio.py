"""Functions for reading/writing Excel from/to 2D array"""
#%%
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook as load_xlsx
from xlrd import open_workbook as load_xls, xldate_as_tuple


def read_xl() -> dict:
    pass

def read_xls(fp) -> Workbook:
    wb = load_xls(fp)
    # wb.

def read_xlsx(fp) -> Workbook:
    return load_xlsx(filename=fp, read_only=True, 
                     keep_vba=False, data_only=True, keep_links=True)


def fromArrays(data) -> Workbook:
    wb = Workbook()
    # Multiple worksheets
    if isinstance(data, dict):
        for nm, tbl in data.items():
            wb.create_sheet(title=nm)
            for row in tbl: wb[nm].append(row)
    # One worksheet
    else:
        ws = wb.active
        for row in tbl: ws.append(row)
    return wb


#%%
from xlrd import open_workbook as load_xls
from xlrd import xldate

def floatHourToTime(fh):
    hours, hourSeconds = divmod(fh, 1)
    minutes, seconds = divmod(hourSeconds * 60, 1)
    return (
        int(hours),
        int(minutes),
        int(seconds * 60),
    )

def xldate2time(excel_date, datemode):
    return xldate.xldate_as_datetime(excel_date, datemode)

def xlrd2pytp(cell, datemode):
    """https://xlrd.readthedocs.io/en/latest/api.html#xlrd.sheet.Cell
    """
    tp = cell.ctype
    val = cell.value

    try:
        if tp == 0: return ""   # empty
        if tp == 1: return val  # string
        if tp == 2: return val  # float
        if tp == 3: return xldate2time(val, datemode)
        if tp == 4: return bool(val)
    except:
        print("error: ")
        print(tp, val)
    return val


wb = load_xls("test.xls", formatting_info=False)
ws = wb[0]
rows = []
for row in ws.get_rows():
    rows.append([ xlrd2pytp(x, wb.datemode) for x in row ])


# rows = list(ws.get_rows())
# cell1 = rows[0][0]
# ce112 = rows[0][2]
# %%
import pandas as pd

d = pd.read_excel("test.xls",  header=None)
# %%
